from datetime import datetime

import openpyxl

from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from .models import Report
from .serializers import ReportSerializer


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all().order_by("-report_date", "-created_at")
    serializer_class = ReportSerializer
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = super().get_queryset()
        request = self.request

        search = request.query_params.get("search", "").strip()
        status_filter = request.query_params.get("status", "").strip()

        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(report_type__icontains=search)
                | Q(category__icontains=search)
                | Q(status__icontains=search)
                | Q(summary__icontains=search)
                | Q(description__icontains=search)
                | Q(prepared_by__icontains=search)
            )

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop("partial", False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except Exception as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        queryset = self.get_queryset()
        return Response({
            "total_reports": queryset.count(),
            "published": queryset.filter(status="published").count(),
            "draft": queryset.filter(status="draft").count(),
            "archived": queryset.filter(status="archived").count(),
        })

    @action(detail=False, methods=["get"], url_path="download-template")
    def download_template(self, request):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reports Template"

            headers = [
                "title",
                "report_type",
                "category",
                "status",
                "summary",
                "description",
                "prepared_by",
                "report_date",
            ]
            sheet.append(headers)

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="reports_template.xlsx"'
            workbook.save(response)
            return response
        except Exception as exc:
            return Response(
                {"detail": f"Template generation failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="upload-excel")
    def upload_excel(self, request):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return Response({"detail": "excel_file is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                return Response({"detail": "Excel file is empty."}, status=status.HTTP_400_BAD_REQUEST)

            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            required_headers = [
                "title",
                "report_type",
                "category",
                "status",
                "summary",
                "description",
                "prepared_by",
                "report_date",
            ]

            missing = [h for h in required_headers if h not in headers]
            if missing:
                return Response(
                    {"detail": f"Missing headers: {', '.join(missing)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            created_count = 0
            errors = []

            for row_index, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue

                row_data = {headers[i]: row[i] for i in range(len(headers))}

                try:
                    title = str(row_data.get("title") or "").strip()
                    report_type = str(row_data.get("report_type") or "other").strip().lower()
                    category = str(row_data.get("category") or "").strip()
                    status_value = str(row_data.get("status") or "draft").strip().lower()
                    summary = str(row_data.get("summary") or "").strip()
                    description = str(row_data.get("description") or "").strip()
                    prepared_by = str(row_data.get("prepared_by") or "").strip()
                    report_date = row_data.get("report_date")

                    if not title:
                        errors.append(f"Row {row_index}: title is required.")
                        continue

                    if report_type not in dict(Report.TYPE_CHOICES):
                        report_type = "other"

                    if status_value not in dict(Report.STATUS_CHOICES):
                        status_value = "draft"

                    if isinstance(report_date, datetime):
                        report_date = report_date.date()
                    elif isinstance(report_date, str) and report_date.strip():
                        report_date = datetime.strptime(report_date.strip(), "%Y-%m-%d").date()
                    else:
                        report_date = timezone.now().date()

                    Report.objects.create(
                        title=title,
                        report_type=report_type,
                        category=category,
                        status=status_value,
                        summary=summary,
                        description=description,
                        prepared_by=prepared_by,
                        report_date=report_date,
                    )
                    created_count += 1

                except Exception as inner_exc:
                    errors.append(f"Row {row_index}: {str(inner_exc)}")

            return Response({
                "message": "Excel upload completed.",
                "created_count": created_count,
                "errors": errors,
            })

        except Exception as exc:
            return Response(
                {"detail": f"Excel upload failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            queryset = self.get_queryset()

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reports"

            headers = [
                "ID",
                "Title",
                "Type",
                "Category",
                "Status",
                "Summary",
                "Description",
                "Prepared By",
                "Report Date",
                "Published At",
                "Created At",
            ]
            sheet.append(headers)

            for obj in queryset:
                sheet.append([
                    obj.id,
                    obj.title,
                    obj.get_report_type_display(),
                    obj.category,
                    obj.get_status_display(),
                    obj.summary,
                    obj.description,
                    obj.prepared_by,
                    obj.report_date.strftime("%Y-%m-%d") if obj.report_date else "",
                    obj.published_at.strftime("%Y-%m-%d %H:%M:%S") if obj.published_at else "",
                    obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else "",
                ])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="reports.xlsx"'
            workbook.save(response)
            return response

        except Exception as exc:
            return Response(
                {"detail": f"Excel export failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        try:
            queryset = self.get_queryset()
            from weasyprint import HTML

            html_string = render_to_string(
                "reports/reports_pdf.html",
                {
                    "reports": queryset,
                    "generated_at": timezone.now(),
                },
            )

            pdf_file = HTML(string=html_string).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="reports.pdf"'
            return response

        except Exception as exc:
            return Response(
                {"detail": f"PDF export failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )