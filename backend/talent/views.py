import io
from datetime import datetime

import openpyxl
from django.db.models import Count
from django.http import FileResponse, HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response

from .models import TalentProfile
from .serializers import TalentProfileSerializer


class TalentProfileViewSet(viewsets.ModelViewSet):
    queryset = TalentProfile.objects.all().order_by("-created_at", "-id")
    serializer_class = TalentProfileSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        "registration_no",
        "student_name",
        "class_name",
        "section",
        "sport",
        "event_or_position",
        "phone",
        "email",
        "guardian_name",
    ]
    ordering_fields = [
        "registration_no",
        "student_name",
        "date_of_birth",
        "class_name",
        "section",
        "sport",
        "talent_level",
        "status",
        "created_at",
    ]
    ordering = ["-created_at", "-id"]

    def get_queryset(self):
        queryset = super().get_queryset()

        gender = self.request.query_params.get("gender")
        level = self.request.query_params.get("talent_level")
        status_value = self.request.query_params.get("status")
        class_name = self.request.query_params.get("class_name")
        section = self.request.query_params.get("section")
        sport = self.request.query_params.get("sport")

        if gender:
            queryset = queryset.filter(gender=gender)

        if level:
            queryset = queryset.filter(talent_level=level)

        if status_value:
            queryset = queryset.filter(status=status_value)

        if class_name:
            queryset = queryset.filter(class_name__iexact=class_name.strip())

        if section:
            queryset = queryset.filter(section__iexact=section.strip())

        if sport:
            queryset = queryset.filter(sport__icontains=sport.strip())

        return queryset

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        queryset = self.get_queryset()

        return Response(
            {
                "total_profiles": queryset.count(),
                "active_profiles": queryset.filter(status="active").count(),
                "selected_profiles": queryset.filter(status="selected").count(),
                "male_count": queryset.filter(gender="M").count(),
                "female_count": queryset.filter(gender="F").count(),
                "sport_distribution": list(
                    queryset.values("sport").annotate(count=Count("id")).order_by("sport")
                ),
                "level_distribution": list(
                    queryset.values("talent_level")
                    .annotate(count=Count("id"))
                    .order_by("talent_level")
                ),
                "status_distribution": list(
                    queryset.values("status").annotate(count=Count("id")).order_by("status")
                ),
            }
        )

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "TalentTemplate"

        headers = [
            "student_name",
            "gender",
            "date_of_birth",
            "class_name",
            "section",
            "sport",
            "event_or_position",
            "talent_level",
            "phone",
            "email",
            "address",
            "guardian_name",
            "guardian_phone",
            "height_cm",
            "weight_kg",
            "blood_group",
            "medical_notes",
            "previous_achievements",
            "notes",
            "status",
        ]
        sheet.append(headers)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="talent_registry_template.xlsx"'
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="upload",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_excel(self, request):
        excel_file = request.FILES.get("file") or request.FILES.get("excel_file")
        if not excel_file:
            return Response(
                {"detail": "Excel file is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response(
                    {"detail": "Excel file is empty."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]

            required_headers = ["student_name", "gender", "sport"]
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                return Response(
                    {"detail": f"Missing required columns: {', '.join(missing_headers)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            allowed_genders = {"M", "F"}
            allowed_levels = {"beginner", "intermediate", "advanced", "elite"}
            allowed_statuses = {"active", "inactive", "selected", "rejected"}

            created_count = 0
            errors = []

            def get_value(row_data, column_name, default=""):
                if column_name not in headers:
                    return default
                idx = headers.index(column_name)
                if idx >= len(row_data):
                    return default
                value = row_data[idx]
                if value is None:
                    return default
                return value

            def clean_text(value, default=""):
                if value is None:
                    return default
                return str(value).strip()

            def normalize_gender(value):
                text = clean_text(value).upper()
                if text in {"M", "MALE"}:
                    return "M"
                if text in {"F", "FEMALE"}:
                    return "F"
                return text

            def normalize_level(value):
                text = clean_text(value, "beginner").lower()
                mapping = {
                    "beginner": "beginner",
                    "intermediate": "intermediate",
                    "advanced": "advanced",
                    "elite": "elite",
                }
                return mapping.get(text, text)

            def normalize_status(value):
                text = clean_text(value, "active").lower()
                mapping = {
                    "active": "active",
                    "inactive": "inactive",
                    "selected": "selected",
                    "rejected": "rejected",
                }
                return mapping.get(text, text)

            def parse_date(value):
                if value in ("", None):
                    return None

                if hasattr(value, "date"):
                    try:
                        return value.date()
                    except Exception:
                        pass

                text = str(value).strip()
                if not text:
                    return None

                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        return datetime.strptime(text, fmt).date()
                    except ValueError:
                        continue
                raise ValueError("invalid date format")

            def parse_float(value):
                if value in ("", None):
                    return None

                if isinstance(value, (int, float)):
                    return float(value)

                text = str(value).strip().replace(",", "")
                if not text:
                    return None
                return float(text)

            for row_number, row_data in enumerate(rows[1:], start=2):
                try:
                    if not row_data or not any(cell not in ("", None) for cell in row_data):
                        continue

                    student_name = clean_text(get_value(row_data, "student_name"))
                    gender = normalize_gender(get_value(row_data, "gender"))
                    sport = clean_text(get_value(row_data, "sport"))

                    if not student_name:
                        errors.append(f"Row {row_number}: student_name is required.")
                        continue

                    if gender not in allowed_genders:
                        errors.append(f"Row {row_number}: gender must be M or F.")
                        continue

                    if not sport:
                        errors.append(f"Row {row_number}: sport is required.")
                        continue

                    talent_level = normalize_level(
                        get_value(row_data, "talent_level", "beginner")
                    )
                    if talent_level not in allowed_levels:
                        errors.append(
                            f"Row {row_number}: talent_level must be one of {', '.join(sorted(allowed_levels))}."
                        )
                        continue

                    status_value = normalize_status(
                        get_value(row_data, "status", "active")
                    )
                    if status_value not in allowed_statuses:
                        errors.append(
                            f"Row {row_number}: status must be one of {', '.join(sorted(allowed_statuses))}."
                        )
                        continue

                    try:
                        date_of_birth = parse_date(get_value(row_data, "date_of_birth"))
                    except ValueError:
                        errors.append(
                            f"Row {row_number}: invalid date_of_birth. Use YYYY-MM-DD or DD/MM/YYYY."
                        )
                        continue

                    try:
                        height_cm = parse_float(get_value(row_data, "height_cm"))
                        weight_kg = parse_float(get_value(row_data, "weight_kg"))
                    except ValueError:
                        errors.append(
                            f"Row {row_number}: height_cm / weight_kg must be numeric."
                        )
                        continue

                    TalentProfile.objects.create(
                        student_name=student_name,
                        gender=gender,
                        date_of_birth=date_of_birth,
                        class_name=clean_text(get_value(row_data, "class_name")),
                        section=clean_text(get_value(row_data, "section")),
                        sport=sport,
                        event_or_position=clean_text(get_value(row_data, "event_or_position")),
                        talent_level=talent_level,
                        phone=clean_text(get_value(row_data, "phone")),
                        email=clean_text(get_value(row_data, "email")),
                        address=clean_text(get_value(row_data, "address")),
                        guardian_name=clean_text(get_value(row_data, "guardian_name")),
                        guardian_phone=clean_text(get_value(row_data, "guardian_phone")),
                        height_cm=height_cm,
                        weight_kg=weight_kg,
                        blood_group=clean_text(get_value(row_data, "blood_group")),
                        medical_notes=clean_text(get_value(row_data, "medical_notes")),
                        previous_achievements=clean_text(
                            get_value(row_data, "previous_achievements")
                        ),
                        notes=clean_text(get_value(row_data, "notes")),
                        status=status_value,
                    )
                    created_count += 1

                except Exception as exc:
                    errors.append(f"Row {row_number}: {str(exc)}")

            return Response(
                {
                    "message": "Excel processed successfully.",
                    "created_count": created_count,
                    "error_count": len(errors),
                    "errors": errors,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as exc:
            return Response(
                {"detail": f"Excel upload failed: {str(exc)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.get_queryset()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "TalentProfiles"

        headers = [
            "registration_no",
            "student_name",
            "gender",
            "date_of_birth",
            "class_name",
            "section",
            "sport",
            "event_or_position",
            "talent_level",
            "phone",
            "email",
            "guardian_name",
            "guardian_phone",
            "blood_group",
            "status",
        ]
        sheet.append(headers)

        for obj in queryset:
            sheet.append(
                [
                    obj.registration_no,
                    obj.student_name,
                    obj.gender,
                    obj.date_of_birth.isoformat() if obj.date_of_birth else "",
                    obj.class_name,
                    obj.section,
                    obj.sport,
                    obj.event_or_position,
                    obj.talent_level,
                    obj.phone,
                    obj.email,
                    obj.guardian_name,
                    obj.guardian_phone,
                    obj.blood_group,
                    obj.status,
                ]
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="talent_registry_export.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        queryset = self.get_queryset()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph("Talent Registry Report", styles["Title"]))
        elements.append(Spacer(1, 0.2 * inch))

        data = [
            [
                "Reg No",
                "Name",
                "Gender",
                "Class",
                "Section",
                "Sport",
                "Level",
                "Phone",
                "Status",
            ]
        ]

        for obj in queryset:
            data.append(
                [
                    obj.registration_no or "",
                    obj.student_name or "",
                    obj.get_gender_display() if obj.gender else "",
                    obj.class_name or "",
                    obj.section or "",
                    obj.sport or "",
                    obj.get_talent_level_display() if obj.talent_level else "",
                    obj.phone or "",
                    obj.get_status_display() if obj.status else "",
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbe7ff")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#173f9f")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbff")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return FileResponse(
            buffer,
            as_attachment=True,
            filename="talent_registry_report.pdf",
        )