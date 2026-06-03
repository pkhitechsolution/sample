from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import datetime

import openpyxl
from openpyxl import Workbook

from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from weasyprint import HTML

from .models import Grant
from .serializers import GrantSerializer


def _parse_decimal(value, default="0"):
    if value in [None, ""]:
        return Decimal(default)
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _parse_excel_date(value):
    if value in [None, ""]:
        return None

    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        parsed = parse_date(value)
        if parsed:
            return parsed

    return None


class GrantViewSet(viewsets.ModelViewSet):
    queryset = Grant.objects.all()
    serializer_class = GrantSerializer
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = Grant.objects.all().order_by("-applied_date", "-id")

        search = self.request.query_params.get("search", "").strip()
        status_filter = self.request.query_params.get("status", "").strip()
        grant_type = self.request.query_params.get("grant_type", "").strip()
        funding_agency = self.request.query_params.get("funding_agency", "").strip()

        if search:
            queryset = queryset.filter(
                Q(grant_name__icontains=search)
                | Q(grant_type__icontains=search)
                | Q(funding_agency__icontains=search)
                | Q(purpose__icontains=search)
                | Q(remarks__icontains=search)
                | Q(status__icontains=search)
            )

        if status_filter:
            queryset = queryset.filter(status__iexact=status_filter)

        if grant_type:
            queryset = queryset.filter(grant_type__iexact=grant_type)

        if funding_agency:
            queryset = queryset.filter(funding_agency__iexact=funding_agency)

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        queryset = Grant.objects.all()

        totals = queryset.aggregate(
            total_requested=Sum("amount_requested"),
            total_approved=Sum("amount_approved"),
        )

        return Response({
            "total_grants": queryset.count(),
            "approved": queryset.filter(status__iexact="approved").count(),
            "pending": queryset.filter(status__iexact="pending").count(),
            "rejected": queryset.filter(status__iexact="rejected").count(),
            "draft": queryset.filter(status__iexact="draft").count(),
            "total_requested": totals["total_requested"] or 0,
            "total_approved": totals["total_approved"] or 0,
        })

    @action(detail=False, methods=["get"], url_path="download-template")
    def download_template(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "GrantsTemplate"

        headers = [
            "grant_name",
            "grant_type",
            "funding_agency",
            "amount_requested",
            "amount_approved",
            "status",
            "applied_date",
            "approval_date",
            "purpose",
            "remarks",
        ]
        sheet.append(headers)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="grants_template.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Grants"

        headers = [
            "grant_name",
            "grant_type",
            "funding_agency",
            "amount_requested",
            "amount_approved",
            "status",
            "applied_date",
            "approval_date",
            "purpose",
            "remarks",
        ]
        sheet.append(headers)

        queryset = Grant.objects.all().order_by("-applied_date", "-id")
        for item in queryset:
            sheet.append([
                item.grant_name or "",
                item.grant_type or "",
                item.funding_agency or "",
                float(item.amount_requested or 0),
                float(item.amount_approved or 0),
                item.status or "",
                item.applied_date.strftime("%Y-%m-%d") if item.applied_date else "",
                item.approval_date.strftime("%Y-%m-%d") if item.approval_date else "",
                item.purpose or "",
                item.remarks or "",
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="grants_export.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        try:
            queryset = Grant.objects.all().order_by("-applied_date", "-id")

            rows_html = ""
            for index, item in enumerate(queryset, start=1):
                rows_html += f"""
                    <tr>
                        <td>{index}</td>
                        <td>{item.grant_name or ""}</td>
                        <td>{item.grant_type or ""}</td>
                        <td>{item.funding_agency or ""}</td>
                        <td>{item.amount_requested or 0}</td>
                        <td>{item.amount_approved or 0}</td>
                        <td>{item.status or ""}</td>
                        <td>{item.applied_date.strftime("%d-%m-%Y") if item.applied_date else ""}</td>
                    </tr>
                """

            if not rows_html:
                rows_html = """
                    <tr>
                        <td colspan="8">No grants found.</td>
                    </tr>
                """

            html_string = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Grants Report</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        font-size: 12px;
                        color: #222;
                        padding: 20px;
                    }}
                    h2 {{
                        margin-bottom: 14px;
                        color: #173b8f;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                    }}
                    th, td {{
                        border: 1px solid #cfcfcf;
                        padding: 8px;
                        text-align: left;
                        vertical-align: top;
                    }}
                    th {{
                        background: #f2f4f8;
                    }}
                </style>
            </head>
            <body>
                <h2>Grants Report</h2>
                <table>
                    <thead>
                        <tr>
                            <th>S.No</th>
                            <th>Grant Name</th>
                            <th>Type</th>
                            <th>Funding Agency</th>
                            <th>Requested</th>
                            <th>Approved</th>
                            <th>Status</th>
                            <th>Applied Date</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows_html}
                    </tbody>
                </table>
            </body>
            </html>
            """

            pdf_file = HTML(string=html_string).write_pdf()

            response = HttpResponse(pdf_file, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="grants_report.pdf"'
            return response

        except Exception as exc:
            return Response(
                {"detail": f"PDF export failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(
        detail=False,
        methods=["post"],
        url_path="upload-excel",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_excel(self, request):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return Response(
                {"detail": "excel_file is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return Response(
                {"detail": "Excel file is empty."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        normalized_headers = [h.lower() for h in headers]

        required_headers = [
            "grant_name",
            "grant_type",
            "funding_agency",
            "amount_requested",
            "amount_approved",
            "status",
            "applied_date",
            "approval_date",
            "purpose",
            "remarks",
        ]

        missing_headers = [h for h in required_headers if h not in normalized_headers]
        if missing_headers:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing_headers)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0

        for row in rows[1:]:
            if not row or all(cell in [None, ""] for cell in row):
                continue

            row_data = dict(zip(normalized_headers, row))

            Grant.objects.create(
                grant_name=str(row_data.get("grant_name") or "").strip(),
                grant_type=str(row_data.get("grant_type") or "other").strip().lower(),
                funding_agency=str(row_data.get("funding_agency") or "").strip(),
                amount_requested=_parse_decimal(row_data.get("amount_requested")),
                amount_approved=_parse_decimal(row_data.get("amount_approved")),
                status=str(row_data.get("status") or "pending").strip().lower(),
                applied_date=_parse_excel_date(row_data.get("applied_date")),
                approval_date=_parse_excel_date(row_data.get("approval_date")),
                purpose=str(row_data.get("purpose") or "").strip(),
                remarks=str(row_data.get("remarks") or "").strip(),
            )
            created_count += 1

        return Response(
            {
                "detail": "Excel uploaded successfully.",
                "created": created_count,
            },
            status=status.HTTP_201_CREATED,
        )