import io
from datetime import datetime

from django.http import HttpResponse
from django.utils.dateparse import parse_date, parse_time

from rest_framework import status, viewsets, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Match
from .serializers import MatchSerializer


class MatchViewSet(viewsets.ModelViewSet):
    queryset = Match.objects.all().order_by("-id")
    serializer_class = MatchSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        "title",
        "tournament_name",
        "sport_name",
        "team_a",
        "team_b",
        "venue",
        "status",
        "match_type",
        "round_name",
        "referee_name",
        "winner",
    ]
    ordering_fields = [
        "id",
        "title",
        "match_date",
        "match_time",
        "status",
        "match_type",
        "venue",
    ]


@api_view(["GET"])
@permission_classes([AllowAny])
def matches_dashboard_summary(request):
    qs = Match.objects.all()

    total_matches = qs.count()
    scheduled_count = qs.filter(status__iexact="scheduled").count()
    ongoing_count = qs.filter(status__in=["ongoing", "live"]).count()
    completed_count = qs.filter(status__iexact="completed").count()
    cancelled_count = qs.filter(status__in=["cancelled", "canceled", "postponed"]).count()

    return Response({
        "total_matches": total_matches,
        "scheduled_count": scheduled_count,
        "ongoing_count": ongoing_count,
        "completed_count": completed_count,
        "cancelled_count": cancelled_count,
    })


@api_view(["GET"])
@permission_classes([AllowAny])
def download_matches_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches Template"

    headers = [
        "Title",
        "Tournament Name",
        "Sport Name",
        "Team A",
        "Team B",
        "Venue",
        "Match Date",
        "Match Time",
        "Match Type",
        "Status",
        "Round Name",
        "Referee Name",
        "Score Team A",
        "Score Team B",
        "Winner",
        "Notes",
    ]
    ws.append(headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="matches_template.xlsx"'
    return response


@api_view(["POST"])
@permission_classes([AllowAny])
def upload_matches_excel(request):
    excel_file = request.FILES.get("file") or request.FILES.get("excel_file")
    if not excel_file:
        return Response(
            {"detail": "Excel file is required."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception:
        return Response(
            {"detail": "Invalid Excel file."},
            status=status.HTTP_400_BAD_REQUEST
        )

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response(
            {"detail": "Excel file is empty."},
            status=status.HTTP_400_BAD_REQUEST
        )

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def row_dict(row):
        values = list(row) + [None] * (len(headers) - len(row))
        return {headers[i]: values[i] for i in range(len(headers))}

    created_count = 0
    updated_count = 0
    skipped_rows = []

    for index, row in enumerate(rows[1:], start=2):
        item = row_dict(row)

        title = str(item.get("Title") or "").strip()
        tournament_name = str(item.get("Tournament Name") or "").strip()
        sport_name = str(item.get("Sport Name") or "").strip()
        team_a = str(item.get("Team A") or "").strip()
        team_b = str(item.get("Team B") or "").strip()
        venue = str(item.get("Venue") or "").strip()
        match_type = str(item.get("Match Type") or "league").strip().lower()
        status_value = str(item.get("Status") or "scheduled").strip().lower()
        round_name = str(item.get("Round Name") or "").strip()
        referee_name = str(item.get("Referee Name") or "").strip()
        winner = str(item.get("Winner") or "").strip()
        notes = str(item.get("Notes") or "").strip()

        score_team_a = item.get("Score Team A") or 0
        score_team_b = item.get("Score Team B") or 0

        match_date_raw = item.get("Match Date")
        match_time_raw = item.get("Match Time")

        if not title or not sport_name or not team_a or not team_b:
            skipped_rows.append({
                "row": index,
                "reason": "Missing required fields"
            })
            continue

        match_date = None
        if match_date_raw:
            if isinstance(match_date_raw, datetime):
                match_date = match_date_raw.date()
            else:
                match_date = parse_date(str(match_date_raw))

        match_time = None
        if match_time_raw:
            if isinstance(match_time_raw, datetime):
                match_time = match_time_raw.time()
            else:
                match_time = parse_time(str(match_time_raw))

        defaults = {
            "tournament_name": tournament_name,
            "sport_name": sport_name,
            "team_a": team_a,
            "team_b": team_b,
            "venue": venue,
            "match_date": match_date,
            "match_time": match_time,
            "match_type": match_type,
            "status": status_value,
            "round_name": round_name,
            "referee_name": referee_name,
            "score_team_a": int(score_team_a or 0),
            "score_team_b": int(score_team_b or 0),
            "winner": winner,
            "notes": notes,
        }

        _, created = Match.objects.update_or_create(
            title=title,
            defaults=defaults,
        )

        if created:
            created_count += 1
        else:
            updated_count += 1

    return Response({
        "detail": "Matches Excel uploaded successfully.",
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": len(skipped_rows),
        "skipped_rows": skipped_rows,
    })


@api_view(["GET"])
@permission_classes([AllowAny])
def export_matches_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"

    headers = [
        "Title",
        "Tournament Name",
        "Sport Name",
        "Team A",
        "Team B",
        "Venue",
        "Match Date",
        "Match Time",
        "Match Type",
        "Status",
        "Round Name",
        "Referee Name",
        "Score Team A",
        "Score Team B",
        "Winner",
        "Notes",
    ]
    ws.append(headers)

    for item in Match.objects.all().order_by("-id"):
        ws.append([
            item.title,
            item.tournament_name,
            item.sport_name,
            item.team_a,
            item.team_b,
            item.venue,
            item.match_date.strftime("%Y-%m-%d") if item.match_date else "",
            item.match_time.strftime("%H:%M:%S") if item.match_time else "",
            item.match_type,
            item.status,
            item.round_name,
            item.referee_name,
            item.score_team_a,
            item.score_team_b,
            item.winner,
            item.notes,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="matches_export.xlsx"'
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def export_matches_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("Matches Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [[
        "S.No",
        "Title",
        "Tournament",
        "Sport",
        "Team A",
        "Team B",
        "Venue",
        "Date",
        "Status",
    ]]

    queryset = Match.objects.all().order_by("-id")
    for index, item in enumerate(queryset, start=1):
        data.append([
            str(index),
            item.title or "",
            item.tournament_name or "",
            item.sport_name or "",
            item.team_a or "",
            item.team_b or "",
            item.venue or "",
            item.match_date.strftime("%d-%m-%Y") if item.match_date else "",
            item.status or "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21489e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="matches_report.pdf"'
    return response