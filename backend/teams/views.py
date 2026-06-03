import io

from django.db.models import F
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status, viewsets
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import Team
from .serializers import TeamSerializer


class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all().order_by("-id")
    serializer_class = TeamSerializer


@api_view(["GET"])
def teams_summary(request):
    total_teams = Team.objects.count()
    active_teams = Team.objects.filter(status__iexact="active").count()
    inactive_teams = Team.objects.filter(status__iexact="inactive").count()
    full_teams = Team.objects.filter(current_players_count__gte=F("max_players")).count()

    return Response(
        {
            "total_teams": total_teams,
            "active_teams": active_teams,
            "inactive_teams": inactive_teams,
            "full_teams": full_teams,
        }
    )


@api_view(["GET"])
def download_teams_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams Template"

    headers = [
        "team_name",
        "sport_name",
        "age_group",
        "gender_category",
        "coach_name",
        "captain_name",
        "vice_captain_name",
        "max_players",
        "current_players_count",
        "academic_year",
        "status",
        "achievements",
        "notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1E3F98")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="teams_template.xlsx"'
    return response


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def upload_teams_excel(request):
    file = request.FILES.get("file")
    if not file:
        return Response(
            {"detail": "No file uploaded."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        workbook = load_workbook(file)
        sheet = workbook.active

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

        required_headers = [
            "team_name",
            "sport_name",
            "age_group",
            "gender_category",
            "coach_name",
            "captain_name",
            "vice_captain_name",
            "max_players",
            "current_players_count",
            "academic_year",
            "status",
            "achievements",
            "notes",
        ]

        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            return Response(
                {
                    "detail": f"Missing required columns: {', '.join(missing_headers)}"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        updated_count = 0
        skipped_rows = []

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue

            row_data = dict(zip(headers, row))

            team_name = str(row_data.get("team_name") or "").strip()
            sport_name = str(row_data.get("sport_name") or "").strip()

            if not team_name or not sport_name:
                skipped_rows.append(
                    f"Row {row_index}: team_name and sport_name are required."
                )
                continue

            max_players_raw = row_data.get("max_players")
            current_players_raw = row_data.get("current_players_count")

            try:
                max_players = int(max_players_raw or 0)
            except (TypeError, ValueError):
                skipped_rows.append(f"Row {row_index}: invalid max_players.")
                continue

            try:
                current_players_count = int(current_players_raw or 0)
            except (TypeError, ValueError):
                skipped_rows.append(f"Row {row_index}: invalid current_players_count.")
                continue

            if max_players <= 0:
                skipped_rows.append(
                    f"Row {row_index}: max_players must be greater than 0."
                )
                continue

            if current_players_count < 0:
                skipped_rows.append(
                    f"Row {row_index}: current_players_count cannot be negative."
                )
                continue

            if current_players_count > max_players:
                skipped_rows.append(
                    f"Row {row_index}: current_players_count cannot exceed max_players."
                )
                continue

            payload = {
                "team_name": team_name,
                "sport_name": sport_name,
                "age_group": str(row_data.get("age_group") or "").strip(),
                "gender_category": str(row_data.get("gender_category") or "Mixed").strip() or "Mixed",
                "coach_name": str(row_data.get("coach_name") or "").strip(),
                "captain_name": str(row_data.get("captain_name") or "").strip(),
                "vice_captain_name": str(row_data.get("vice_captain_name") or "").strip(),
                "max_players": max_players,
                "current_players_count": current_players_count,
                "academic_year": str(row_data.get("academic_year") or "").strip(),
                "status": str(row_data.get("status") or "Active").strip() or "Active",
                "achievements": str(row_data.get("achievements") or "").strip(),
                "notes": str(row_data.get("notes") or "").strip(),
            }

            obj, created = Team.objects.update_or_create(
                team_name=team_name,
                sport_name=sport_name,
                defaults=payload,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        return Response(
            {
                "detail": "Teams Excel processed successfully.",
                "created_count": created_count,
                "updated_count": updated_count,
                "skipped_rows": skipped_rows,
            },
            status=status.HTTP_200_OK,
        )

    except Exception as exc:
        return Response(
            {"detail": f"Excel upload failed: {str(exc)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["GET"])
def export_teams_excel(request):
    teams = Team.objects.all().order_by("team_name", "sport_name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Teams Export"

    headers = [
        "S.No",
        "team_name",
        "sport_name",
        "age_group",
        "gender_category",
        "coach_name",
        "captain_name",
        "vice_captain_name",
        "max_players",
        "current_players_count",
        "vacancies",
        "fill_percentage",
        "academic_year",
        "status",
        "achievements",
        "notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1E3F98")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for index, team in enumerate(teams, start=1):
        max_players = team.max_players or 0
        current_players = team.current_players_count or 0
        vacancies = max(max_players - current_players, 0)
        fill_percentage = round((current_players / max_players) * 100, 2) if max_players > 0 else 0

        ws.append(
            [
                index,
                team.team_name,
                team.sport_name,
                team.age_group,
                team.gender_category,
                team.coach_name,
                team.captain_name,
                team.vice_captain_name,
                max_players,
                current_players,
                vacancies,
                fill_percentage,
                team.academic_year,
                team.status,
                team.achievements,
                team.notes,
            ]
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="teams_export.xlsx"'
    return response


@api_view(["GET"])
def export_teams_pdf(request):
    teams = Team.objects.all().order_by("team_name", "sport_name")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Teams Report", styles["Title"])
    subtitle = Paragraph(
        f"Generated on: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}",
        styles["Normal"],
    )

    story.append(title)
    story.append(Spacer(1, 0.15 * inch))
    story.append(subtitle)
    story.append(Spacer(1, 0.25 * inch))

    table_data = [
        [
            "S.No",
            "Team",
            "Sport",
            "Age Group",
            "Category",
            "Coach",
            "Captain",
            "Players",
            "Vacancies",
            "Status",
        ]
    ]

    for index, team in enumerate(teams, start=1):
        max_players = team.max_players or 0
        current_players = team.current_players_count or 0
        vacancies = max(max_players - current_players, 0)

        table_data.append(
            [
                str(index),
                team.team_name or "",
                team.sport_name or "",
                team.age_group or "",
                team.gender_category or "",
                team.coach_name or "",
                team.captain_name or "",
                f"{current_players}/{max_players}",
                str(vacancies),
                team.status or "",
            ]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3F98")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
            ]
        )
    )

    story.append(table)
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="teams_report.pdf"'
    return response