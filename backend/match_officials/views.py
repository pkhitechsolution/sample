import io
from datetime import datetime

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from .models import Official
from .serializers import OfficialSerializer


class OfficialViewSet(viewsets.ModelViewSet):
    queryset = Official.objects.all().order_by("-id")
    serializer_class = OfficialSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = Official.objects.all().order_by("-id")

        search = self.request.query_params.get("search", "").strip()
        role = self.request.query_params.get("role", "").strip()
        status_value = self.request.query_params.get("status", "").strip()
        availability = self.request.query_params.get("availability", "").strip()
        gender = self.request.query_params.get("gender", "").strip()
        sport = self.request.query_params.get("sport", "").strip()

        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search)
                | Q(official_code__icontains=search)
                | Q(role__icontains=search)
                | Q(sport__icontains=search)
                | Q(phone__icontains=search)
                | Q(email__icontains=search)
                | Q(qualification__icontains=search)
                | Q(city__icontains=search)
            )

        if role:
            queryset = queryset.filter(role__iexact=role)

        if status_value:
            queryset = queryset.filter(status__iexact=status_value)

        if availability:
            queryset = queryset.filter(availability__iexact=availability)

        if gender:
            queryset = queryset.filter(gender__iexact=gender)

        if sport:
            queryset = queryset.filter(sport__icontains=sport)

        return queryset


@api_view(["GET"])
@permission_classes([AllowAny])
def officials_summary(request):
    queryset = Official.objects.all()

    return Response(
        {
            "total_officials": queryset.count(),
            "active_officials": queryset.filter(status="active").count(),
            "inactive_officials": queryset.filter(status="inactive").count(),
            "available_officials": queryset.filter(availability="available").count(),
            "busy_officials": queryset.filter(availability="busy").count(),
            "unavailable_officials": queryset.filter(availability="unavailable").count(),
            "referees": queryset.filter(role="referee").count(),
            "umpires": queryset.filter(role="umpire").count(),
            "judges": queryset.filter(role="judge").count(),
            "marshals": queryset.filter(role="marshal").count(),
            "scorers": queryset.filter(role="scorer").count(),
            "time_keepers": queryset.filter(role="time_keeper").count(),
            "coordinators": queryset.filter(role="coordinator").count(),
        }
    )


def _safe_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_role(value):
    value = _safe_str(value).lower().replace(" ", "_")
    allowed = {
        "referee",
        "umpire",
        "judge",
        "marshal",
        "scorer",
        "time_keeper",
        "coordinator",
        "other",
    }
    return value if value in allowed else "other"


def _normalize_gender(value):
    value = _safe_str(value).lower()
    return value if value in {"male", "female", "other"} else ""


def _normalize_availability(value):
    value = _safe_str(value).lower().replace(" ", "_")
    return value if value in {"available", "busy", "unavailable"} else "available"


def _normalize_status(value):
    value = _safe_str(value).lower()
    return value if value in {"active", "inactive"} else "active"


@api_view(["POST"])
@permission_classes([AllowAny])
def upload_officials_excel(request):
    excel_file = request.FILES.get("file") or request.FILES.get("excel_file")
    if not excel_file:
        return Response(
            {"detail": "Please upload an Excel file using field name 'file' or 'excel_file'."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except Exception:
        return Response(
            {"detail": "Invalid Excel file."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return Response(
            {"detail": "Excel file is empty."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    headers = [_safe_str(h).lower() for h in rows[0]]

    def col_index(*names):
        for name in names:
            if name.lower() in headers:
                return headers.index(name.lower())
        return None

    idx_full_name = col_index("full name", "full_name", "name")
    idx_role = col_index("role")
    idx_sport = col_index("sport")
    idx_gender = col_index("gender")
    idx_phone = col_index("phone", "mobile", "contact")
    idx_email = col_index("email", "mail")
    idx_qualification = col_index("qualification")
    idx_experience = col_index("experience years", "experience_years", "experience")
    idx_city = col_index("city")
    idx_address = col_index("address")
    idx_availability = col_index("availability")
    idx_status = col_index("status")
    idx_notes = col_index("notes", "remarks")
    idx_official_code = col_index("official code", "official_code", "code")

    if idx_full_name is None:
        return Response(
            {"detail": "Excel must contain Full Name column."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created_count = 0
    updated_count = 0
    skipped_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        try:
            full_name = _safe_str(row[idx_full_name]) if idx_full_name is not None else ""
            if not full_name:
                skipped_rows.append(f"Row {row_number}: Full Name is required.")
                continue

            official_code = _safe_str(row[idx_official_code]) if idx_official_code is not None else ""
            role = _normalize_role(row[idx_role]) if idx_role is not None else "referee"
            sport = _safe_str(row[idx_sport]) if idx_sport is not None else ""
            gender = _normalize_gender(row[idx_gender]) if idx_gender is not None else ""
            phone = _safe_str(row[idx_phone]) if idx_phone is not None else ""
            email = _safe_str(row[idx_email]) if idx_email is not None else ""
            qualification = _safe_str(row[idx_qualification]) if idx_qualification is not None else ""
            city = _safe_str(row[idx_city]) if idx_city is not None else ""
            address = _safe_str(row[idx_address]) if idx_address is not None else ""
            availability = _normalize_availability(row[idx_availability]) if idx_availability is not None else "available"
            status_value = _normalize_status(row[idx_status]) if idx_status is not None else "active"
            notes = _safe_str(row[idx_notes]) if idx_notes is not None else ""

            experience_years = 0
            if idx_experience is not None and row[idx_experience] not in (None, ""):
                try:
                    experience_years = int(float(row[idx_experience]))
                except Exception:
                    experience_years = 0

            official = None
            if official_code:
                official = Official.objects.filter(official_code__iexact=official_code).first()
            if official is None and email:
                official = Official.objects.filter(email__iexact=email).first()
            if official is None and phone:
                official = Official.objects.filter(phone__iexact=phone).first()

            payload = {
                "full_name": full_name,
                "role": role,
                "sport": sport,
                "gender": gender,
                "phone": phone,
                "email": email or None,
                "qualification": qualification,
                "experience_years": experience_years,
                "city": city,
                "address": address,
                "availability": availability,
                "status": status_value,
                "notes": notes,
            }

            if official:
                for key, value in payload.items():
                    setattr(official, key, value)
                official.save()

                if official_code and official.official_code != official_code:
                    official.official_code = official_code
                    official.save(update_fields=["official_code"])

                updated_count += 1
            else:
                official = Official.objects.create(**payload)
                if official_code:
                    official.official_code = official_code
                    official.save(update_fields=["official_code"])
                created_count += 1

        except Exception as exc:
            skipped_rows.append(f"Row {row_number}: {str(exc)}")

    return Response(
        {
            "detail": "Officials Excel uploaded successfully.",
            "created_count": created_count,
            "updated_count": updated_count,
            "skipped_count": len(skipped_rows),
            "skipped_rows": skipped_rows,
        },
        status=status.HTTP_200_OK,
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def download_officials_template(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Officials Template"

    sheet.append([
        "Full Name",
        "Official Code",
        "Role",
        "Sport",
        "Gender",
        "Phone",
        "Email",
        "Qualification",
        "Experience Years",
        "City",
        "Address",
        "Availability",
        "Status",
        "Notes",
    ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="officials_template.xlsx"'
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def export_officials_excel(request):
    queryset = Official.objects.all().order_by("-id")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Officials"

    sheet.append([
        "S.No",
        "Full Name",
        "Official Code",
        "Role",
        "Sport",
        "Gender",
        "Phone",
        "Email",
        "Qualification",
        "Experience Years",
        "City",
        "Availability",
        "Status",
        "Notes",
    ])

    for index, official in enumerate(queryset, start=1):
        sheet.append([
            index,
            official.full_name,
            official.official_code or "",
            official.get_role_display(),
            official.sport,
            official.get_gender_display() if official.gender else "",
            official.phone,
            official.email or "",
            official.qualification,
            official.experience_years,
            official.city,
            official.get_availability_display(),
            official.get_status_display(),
            official.notes,
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="officials_export.xlsx"'
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def export_officials_pdf(request):
    queryset = Official.objects.all().order_by("-id")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Officials Report", styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    data = [[
        "S.No",
        "Name",
        "Code",
        "Role",
        "Sport",
        "Phone",
        "Qualification",
        "Experience",
        "Availability",
        "Status",
    ]]

    for index, official in enumerate(queryset, start=1):
        data.append([
            str(index),
            official.full_name or "",
            official.official_code or "",
            official.get_role_display() if official.role else "",
            official.sport or "",
            official.phone or "",
            official.qualification or "",
            str(official.experience_years or 0),
            official.get_availability_display() if official.availability else "",
            official.get_status_display() if official.status else "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3f98")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="officials_report.pdf"'
    return response
