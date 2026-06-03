import io
from datetime import datetime, date

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status, viewsets, filters
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import Tournament
from .serializers import TournamentSerializer


def _parse_excel_date(value):
    if value in [None, ""]:
        return None

    if isinstance(value, date):
        return value

    if isinstance(value, datetime):
        return value.date()

    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    return None


class TournamentViewSet(viewsets.ModelViewSet):
    queryset = Tournament.objects.all().order_by("-id")
    serializer_class = TournamentSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "sport", "venue", "organizer", "contact_person"]
    ordering_fields = [
        "id",
        "name",
        "sport",
        "start_date",
        "end_date",
        "status",
        "teams_count",
        "matches_count",
    ]

    def get_queryset(self):
        queryset = super().get_queryset()

        status_param = self.request.query_params.get("status")
        sport_param = self.request.query_params.get("sport")
        search_param = self.request.query_params.get("search")

        if status_param:
            queryset = queryset.filter(status__iexact=status_param)

        if sport_param:
            queryset = queryset.filter(sport__icontains=sport_param)

        if search_param:
            queryset = queryset.filter(
                Q(name__icontains=search_param)
                | Q(sport__icontains=search_param)
                | Q(venue__icontains=search_param)
                | Q(organizer__icontains=search_param)
            )

        return queryset


@api_view(["GET"])
def tournaments_dashboard_summary(request):
    total = Tournament.objects.count()
    draft = Tournament.objects.filter(status__iexact="Draft").count()
    open_count = Tournament.objects.filter(status__iexact="Open").count()
    scheduled = Tournament.objects.filter(status__iexact="Scheduled").count()
    ongoing = Tournament.objects.filter(status__iexact="Ongoing").count()
    completed = Tournament.objects.filter(status__iexact="Completed").count()

    return Response(
        {
            "total": total,
            "draft": draft,
            "open": open_count,
            "scheduled": scheduled,
            "ongoing": ongoing,
            "completed": completed,
        }
    )


@api_view(["GET"])
def download_tournament_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tournaments Template"

    headers = [
        "name",
        "sport",
        "format",
        "age_group",
        "gender_category",
        "venue",
        "organizer",
        "contact_person",
        "contact_phone",
        "start_date",
        "end_date",
        "registration_last_date",
        "max_teams",
        "teams_count",
        "matches_count",
        "status",
        "description",
        "rules",
        "notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1E3F98")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="tournaments_template.xlsx"'
    return response


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def upload_tournament_excel(request):
    file = request.FILES.get("file")
    if not file:
        return Response(
            {"detail": "No file uploaded."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        workbook = load_workbook(file)
        sheet = workbook.active

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

        required_headers = [
            "name",
            "sport",
            "format",
            "age_group",
            "gender_category",
            "venue",
            "organizer",
            "contact_person",
            "contact_phone",
            "start_date",
            "end_date",
            "registration_last_date",
            "max_teams",
            "teams_count",
            "matches_count",
            "status",
            "description",
            "rules",
            "notes",
        ]

        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing_headers)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        updated_count = 0
        skipped_rows = []

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue

            row_data = dict(zip(headers, row))

            name = str(row_data.get("name") or "").strip()
            sport = str(row_data.get("sport") or "").strip()

            if not name or not sport:
                skipped_rows.append(f"Row {row_index}: name and sport are required.")
                continue

            try:
                max_teams = int(row_data.get("max_teams") or 0)
            except (TypeError, ValueError):
                skipped_rows.append(f"Row {row_index}: invalid max_teams.")
                continue

            try:
                teams_count = int(row_data.get("teams_count") or 0)
            except (TypeError, ValueError):
                skipped_rows.append(f"Row {row_index}: invalid teams_count.")
                continue

            try:
                matches_count = int(row_data.get("matches_count") or 0)
            except (TypeError, ValueError):
                skipped_rows.append(f"Row {row_index}: invalid matches_count.")
                continue

            if max_teams < 0:
                skipped_rows.append(f"Row {row_index}: max_teams cannot be negative.")
                continue

            if teams_count < 0:
                skipped_rows.append(f"Row {row_index}: teams_count cannot be negative.")
                continue

            if matches_count < 0:
                skipped_rows.append(f"Row {row_index}: matches_count cannot be negative.")
                continue

            if max_teams > 0 and teams_count > max_teams:
                skipped_rows.append(f"Row {row_index}: teams_count cannot exceed max_teams.")
                continue

            start_date = _parse_excel_date(row_data.get("start_date"))
            end_date = _parse_excel_date(row_data.get("end_date"))
            registration_last_date = _parse_excel_date(row_data.get("registration_last_date"))

            if not start_date or not end_date:
                skipped_rows.append(f"Row {row_index}: valid start_date and end_date are required.")
                continue

            if end_date < start_date:
                skipped_rows.append(f"Row {row_index}: end_date cannot be earlier than start_date.")
                continue

            payload = {
                "name": name,
                "sport": sport,
                "format": str(row_data.get("format") or "LEAGUE").strip() or "LEAGUE",
                "age_group": str(row_data.get("age_group") or "").strip(),
                "gender_category": str(row_data.get("gender_category") or "Mixed").strip() or "Mixed",
                "venue": str(row_data.get("venue") or "").strip(),
                "organizer": str(row_data.get("organizer") or "").strip(),
                "contact_person": str(row_data.get("contact_person") or "").strip(),
                "contact_phone": str(row_data.get("contact_phone") or "").strip(),
                "start_date": start_date,
                "end_date": end_date,
                "registration_last_date": registration_last_date,
                "max_teams": max_teams,
                "teams_count": teams_count,
                "matches_count": matches_count,
                "status": str(row_data.get("status") or "Draft").strip() or "Draft",
                "description": str(row_data.get("description") or "").strip(),
                "rules": str(row_data.get("rules") or "").strip(),
                "notes": str(row_data.get("notes") or "").strip(),
            }

            _, created = Tournament.objects.update_or_create(
                name=name,
                sport=sport,
                start_date=start_date,
                defaults=payload,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        return Response(
            {
                "detail": "Tournaments Excel processed successfully.",
                "created_count": created_count,
                "updated_count": updated_count,
                "skipped_rows": skipped_rows,
            },
            status=status.HTTP_200_OK,
        )

    except Exception as exc:
        return Response(
            {"detail": f"Excel upload failed: {str(exc)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["GET"])
def export_tournaments_excel(request):
    tournaments = Tournament.objects.all().order_by("name", "sport", "start_date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tournaments Export"

    headers = [
        "S.No",
        "name",
        "sport",
        "format",
        "age_group",
        "gender_category",
        "venue",
        "organizer",
        "contact_person",
        "contact_phone",
        "start_date",
        "end_date",
        "registration_last_date",
        "max_teams",
        "teams_count",
        "matches_count",
        "vacancies",
        "fill_percentage",
        "duration_days",
        "status",
        "description",
        "rules",
        "notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1E3F98")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for index, tournament in enumerate(tournaments, start=1):
        max_teams = tournament.max_teams or 0
        teams_count = tournament.teams_count or 0
        vacancies = max(max_teams - teams_count, 0)
        fill_percentage = round((teams_count / max_teams) * 100, 2) if max_teams > 0 else 0
        duration_days = (
            (tournament.end_date - tournament.start_date).days + 1
            if tournament.start_date and tournament.end_date
            else 0
        )

        ws.append(
            [
                index,
                tournament.name,
                tournament.sport,
                tournament.format,
                tournament.age_group,
                tournament.gender_category,
                tournament.venue,
                tournament.organizer,
                tournament.contact_person,
                tournament.contact_phone,
                str(tournament.start_date) if tournament.start_date else "",
                str(tournament.end_date) if tournament.end_date else "",
                str(tournament.registration_last_date) if tournament.registration_last_date else "",
                max_teams,
                teams_count,
                tournament.matches_count,
                vacancies,
                fill_percentage,
                duration_days,
                tournament.status,
                tournament.description,
                tournament.rules,
                tournament.notes,
            ]
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 32)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="tournaments_export.xlsx"'
    return response


@api_view(["GET"])
def export_tournaments_pdf(request):
    tournaments = Tournament.objects.all().order_by("name", "sport", "start_date")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Tournaments Report", styles["Title"])
    subtitle = Paragraph(
        f"Generated on: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}",
        styles["Normal"],
    )

    story.append(title)
    story.append(Spacer(1, 0.15 * inch))
    story.append(subtitle)
    story.append(Spacer(1, 0.25 * inch))

    table_data = [
        [
            "S.No",
            "Name",
            "Sport",
            "Format",
            "Age Group",
            "Venue",
            "Dates",
            "Teams",
            "Matches",
            "Status",
        ]
    ]

    for index, tournament in enumerate(tournaments, start=1):
        table_data.append(
            [
                str(index),
                tournament.name or "",
                tournament.sport or "",
                tournament.format or "",
                tournament.age_group or "",
                tournament.venue or "",
                f"{tournament.start_date} to {tournament.end_date}",
                f"{tournament.teams_count}/{tournament.max_teams}",
                str(tournament.matches_count or 0),
                tournament.status or "",
            ]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3F98")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
            ]
        )
    )

    story.append(table)
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="tournaments_report.pdf"'
    return response