from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import datetime

import openpyxl
from openpyxl import Workbook

from django.db.models import Avg, Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from weasyprint import HTML

from .models import Performance
from .serializers import PerformanceSerializer


def _parse_decimal(value, default="0"):
    if value in [None, ""]:
        return Decimal(default)
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _parse_excel_date(value):
    if value in [None, ""]:
        return None

    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        parsed = parse_date(value)
        if parsed:
            return parsed

    return None


class PerformanceViewSet(viewsets.ModelViewSet):
    queryset = Performance.objects.all()
    serializer_class = PerformanceSerializer
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = Performance.objects.all().order_by("-performance_date", "-id")

        search = self.request.query_params.get("search", "").strip()
        sport = self.request.query_params.get("sport", "").strip()
        performance_level = self.request.query_params.get("performance_level", "").strip()

        if search:
            queryset = queryset.filter(
                Q(student_name__icontains=search)
                | Q(sport__icontains=search)
                | Q(event_name__icontains=search)
                | Q(coach_name__icontains=search)
                | Q(remarks__icontains=search)
            )

        if sport:
            queryset = queryset.filter(sport__iexact=sport)

        if performance_level:
            queryset = queryset.filter(performance_level__iexact=performance_level)

        return queryset

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        queryset = Performance.objects.all()
        aggregates = queryset.aggregate(avg_score=Avg("performance_score"))

        return Response({
            "total_records": queryset.count(),
            "excellent": queryset.filter(performance_level="excellent").count(),
            "good": queryset.filter(performance_level="good").count(),
            "average": queryset.filter(performance_level="average").count(),
            "needs_improvement": queryset.filter(performance_level="needs_improvement").count(),
            "avg_score": round(float(aggregates["avg_score"] or 0), 2),
        })

    @action(detail=False, methods=["get"], url_path="download-template")
    def download_template(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PerformanceTemplate"

        headers = [
            "student_name",
            "sport",
            "event_name",
            "performance_score",
            "performance_level",
            "performance_date",
            "coach_name",
            "remarks",
        ]
        sheet.append(headers)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="performance_template.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Performance"

        headers = [
            "student_name",
            "sport",
            "event_name",
            "performance_score",
            "performance_level",
            "performance_date",
            "coach_name",
            "remarks",
        ]
        sheet.append(headers)

        queryset = Performance.objects.all().order_by("-performance_date", "-id")
        for item in queryset:
            sheet.append([
                item.student_name or "",
                item.sport or "",
                item.event_name or "",
                float(item.performance_score or 0),
                item.performance_level or "",
                item.performance_date.strftime("%Y-%m-%d") if item.performance_date else "",
                item.coach_name or "",
                item.remarks or "",
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="performance_export.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        try:
            queryset = Performance.objects.all().order_by("-performance_date", "-id")

            rows_html = ""
            for index, item in enumerate(queryset, start=1):
                rows_html += f"""
                    <tr>
                        <td>{index}</td>
                        <td>{item.student_name or ""}</td>
                        <td>{item.sport or ""}</td>
                        <td>{item.event_name or ""}</td>
                        <td>{item.performance_score or 0}</td>
                        <td>{item.performance_level or ""}</td>
                        <td>{item.performance_date.strftime("%d-%m-%Y") if item.performance_date else ""}</td>
                        <td>{item.coach_name or ""}</td>
                    </tr>
                """

            if not rows_html:
                rows_html = """
                    <tr>
                        <td colspan="8">No performance records found.</td>
                    </tr>
                """

            html_string = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Performance Report</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        font-size: 12px;
                        color: #222;
                        padding: 20px;
                    }}
                    h2 {{
                        margin-bottom: 14px;
                        color: #173b8f;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                    }}
                    th, td {{
                        border: 1px solid #cfcfcf;
                        padding: 8px;
                        text-align: left;
                        vertical-align: top;
                    }}
                    th {{
                        background: #f2f4f8;
                    }}
                </style>
            </head>
            <body>
                <h2>Performance Report</h2>
                <table>
                    <thead>
                        <tr>
                            <th>S.No</th>
                            <th>Student Name</th>
                            <th>Sport</th>
                            <th>Event</th>
                            <th>Score</th>
                            <th>Level</th>
                            <th>Date</th>
                            <th>Coach</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows_html}
                    </tbody>
                </table>
            </body>
            </html>
            """

            pdf_file = HTML(string=html_string).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="performance_report.pdf"'
            return response

        except Exception as exc:
            return Response(
                {"detail": f"PDF export failed: {str(exc)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(
        detail=False,
        methods=["post"],
        url_path="upload-excel",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_excel(self, request):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return Response(
                {"detail": "excel_file is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return Response(
                {"detail": "Excel file is empty."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        normalized_headers = [h.lower() for h in headers]

        required_headers = [
            "student_name",
            "sport",
            "event_name",
            "performance_score",
            "performance_level",
            "performance_date",
            "coach_name",
            "remarks",
        ]

        missing_headers = [h for h in required_headers if h not in normalized_headers]
        if missing_headers:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing_headers)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0

        for row in rows[1:]:
            if not row or all(cell in [None, ""] for cell in row):
                continue

            row_data = dict(zip(normalized_headers, row))

            Performance.objects.create(
                student_name=str(row_data.get("student_name") or "").strip(),
                sport=str(row_data.get("sport") or "").strip(),
                event_name=str(row_data.get("event_name") or "").strip(),
                performance_score=_parse_decimal(row_data.get("performance_score")),
                performance_level=str(row_data.get("performance_level") or "average").strip().lower(),
                performance_date=_parse_excel_date(row_data.get("performance_date")),
                coach_name=str(row_data.get("coach_name") or "").strip(),
                remarks=str(row_data.get("remarks") or "").strip(),
            )
            created_count += 1

        return Response(
            {
                "detail": "Excel uploaded successfully.",
                "created": created_count,
            },
            status=status.HTTP_201_CREATED,
        )